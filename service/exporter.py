# service/exporter.py
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime


class DataExporter:
    def __init__(self, db):
        self.db = db

    def export_all_records(self, save_path: str):
        """导出所有面试记录到 Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "面试记录"

        # ===== 表头 =====
        headers = [
            "面试者ID", "姓名", "邮箱", "电话",
            "题目ID", "题目类型", "难度", "题目内容",
            "得分", "备注", "面试时间"
        ]

        ws.append(headers)

        # 表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # ===== 数据查询 =====
        rows = self.db.fetchall("""
                                SELECT i.id,
                                       i.name,
                                       i.email,
                                       i.phone,
                                       r.question_id,
                                       r.score,
                                       r.answer_snapshot,
                                       r.created_at
                                FROM interview_record r
                                         JOIN interviewee i ON r.interviewee_id = i.id
                                ORDER BY r.created_at DESC
                                """)

        # ===== 填充数据 =====
        for row in rows:
            interviewee_id, name, email, phone, q_id, score, snapshot_json, created_at = row

            # 解析 snapshot
            snapshot = json.loads(snapshot_json)
            q_type = snapshot.get("type", "")
            difficulty = snapshot.get("difficulty", "")
            content = snapshot.get("content", "")
            remark = snapshot.get("remark", "")

            ws.append([
                interviewee_id, name, email or "", phone or "",
                q_id, q_type, difficulty, content,
                score, remark, created_at
            ])

        # ===== 列宽调整 =====
        column_widths = [10, 15, 25, 15, 10, 15, 10, 50, 8, 30, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = width

        # ===== 保存 =====
        wb.save(save_path)
        return len(rows)

    def export_interviewee_records(self, interviewee_id: int, save_path: str):
        """导出指定面试者的记录"""
        wb = Workbook()
        ws = wb.active
        ws.title = "个人记录"

        # ===== 面试者信息 =====
        info = self.db.fetchall(
            "SELECT name, email, phone, created_at FROM interviewee WHERE id=?",
            (interviewee_id,)
        )

        if not info:
            raise ValueError(f"未找到面试者 ID={interviewee_id}")

        name, email, phone, created_at = info[0]

        ws.append(["面试者信息"])
        ws.append(["姓名", name])
        ws.append(["邮箱", email or "未填写"])
        ws.append(["电话", phone or "未填写"])
        ws.append(["创建时间", created_at])
        ws.append([])  # 空行

        # ===== 答题记录 =====
        ws.append(["题目类型", "难度", "题目内容", "得分", "备注", "答题时间"])

        rows = self.db.fetchall("""
                                SELECT question_id, score, answer_snapshot, created_at
                                FROM interview_record
                                WHERE interviewee_id = ?
                                ORDER BY created_at
                                """, (interviewee_id,))

        for row in rows:
            q_id, score, snapshot_json, ans_time = row
            snapshot = json.loads(snapshot_json)

            ws.append([
                snapshot.get("type", ""),
                snapshot.get("difficulty", ""),
                snapshot.get("content", ""),
                score,
                snapshot.get("remark", ""),
                ans_time
            ])

        # ===== 列宽调整 =====
        for col in range(1, 7):
            ws.column_dimensions[ws.cell(1, col).column_letter].width = 20

        wb.save(save_path)
        return len(rows)