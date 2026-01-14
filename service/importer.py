# service/importer.py
import csv
from pathlib import Path
from typing import List, Dict

from openpyxl import load_workbook


class QuestionImporter:
    REQUIRED_COLUMNS = {"type", "difficulty", "content", "answer"}

    def __init__(self, db):
        self.db = db

    # ========= 对外统一入口 =========
    def import_from_file(self, file_path: str) -> dict:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(path)

        suffix = path.suffix.lower()

        if suffix == ".csv":
            rows = self._read_csv(path)
        elif suffix in (".xlsx", ".xls"):
            rows = self._read_excel(path)
        else:
            raise ValueError(f"不支持的文件格式: {suffix}")

        if not rows:
            return {"inserted": 0}

        self._insert_rows(rows)
        return {"inserted": len(rows)}

    # ========= CSV =========
    def _read_csv(self, path: Path) -> List[Dict]:
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            headers = set(reader.fieldnames or [])

            self._check_headers(headers)

            return [
                row for row in reader
                if row.get("content")
            ]

    # ========= Excel =========
    def _read_excel(self, path: Path) -> List[Dict]:
        wb = load_workbook(path, read_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(h).strip() for h in rows[0]]
        self._check_headers(set(headers))

        result = []
        for row in rows[1:]:
            data = dict(zip(headers, row))
            if not data.get("content"):
                continue
            result.append(data)

        return result

    # ========= 校验 =========
    def _check_headers(self, headers: set):
        if not self.REQUIRED_COLUMNS.issubset(headers):
            missing = self.REQUIRED_COLUMNS - headers
            raise ValueError(f"缺少字段: {missing}")

    # ========= 入库 =========
    def _insert_rows(self, rows: List[Dict]):
        data = [
            (
                r["type"].strip(),
                r["difficulty"].strip(),
                r["content"].strip(),
                str(r.get("answer", "")).strip()
            )
            for r in rows
        ]

        self.db.executemany("""
            INSERT INTO question_bank
            (q_type, difficulty, content, answer)
            VALUES (?, ?, ?, ?)
        """, data)
